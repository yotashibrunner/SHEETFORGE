"""
SheetForge — deterministic spec -> .xlsx builder.

The LLM writes the SPEC (labels, columns, formula *patterns*, copy).
This code renders the workbook. Formulas are expanded by code, so they
cannot break the way LLM-emitted cells did in v1.

Run:  python forge_build.py specs/reseller.json out.xlsx
(with no args it builds the embedded demo spec)
"""
import sys, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

# ---------- formats ----------
FMT = {
    "currency": '$#,##0.00;($#,##0.00);"–"',
    "currency0": '$#,##0;($#,##0);"–"',
    "percent": '0.0%;-0.0%;"–"',
    "int": '#,##0;;"–"',
    "date": 'mm/dd/yyyy',
    "text": '@',
}
INPUT_BLUE = Font(name="Arial", color="0000CC", size=10)
CALC_BLACK = Font(name="Arial", color="111111", size=10)
LINK_GREEN = Font(name="Arial", color="107C41", size=10)  # cross-sheet
TXT = Font(name="Arial", color="111111", size=10)
THIN = Side(style="thin", color="D9D9E3")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, ncols, accent):
    f = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    fill = PatternFill("solid", fgColor=accent)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = f
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


def build_sheet(ws, sdef, accent):
    cols = sdef["columns"]
    nrows = sdef.get("rows", 100)
    for ci, col in enumerate(cols, start=1):
        ws.cell(row=1, column=ci, value=col["header"])
        ws.column_dimensions[get_column_letter(ci)].width = col.get("width", 16)
    shade = PatternFill("solid", fgColor="F4F3FB")
    for r in range(2, nrows + 2):
        for ci, col in enumerate(cols, start=1):
            cell = ws.cell(row=r, column=ci)
            t = col["type"]
            fmt = col.get("format")
            if t == "input":
                cell.value = 0 if fmt in ("currency", "currency0", "int", "percent") else None
                cell.font = INPUT_BLUE
            elif t == "formula":
                cell.value = col["formula"].replace("{row}", str(r))
                cell.font = LINK_GREEN if col.get("link") else CALC_BLACK
            elif t == "seed":  # pre-filled label cells (e.g. month names)
                cell.value = col["values"][r - 2] if (r - 2) < len(col["values"]) else None
                cell.font = TXT
            else:
                cell.font = TXT
            if fmt in FMT:
                cell.number_format = FMT[fmt]
            cell.border = BORDER
            if r % 2 == 0:
                cell.fill = shade
    if sdef.get("autofilter", True):
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"


def build_dashboard(ws, dash, accent):
    ws.sheet_view.showGridLines = False
    ws["B2"] = dash.get("title", "Dashboard")
    ws["B2"].font = Font(name="Arial", bold=True, size=22, color="1A1A2E")
    ws["B3"] = dash.get("subtitle", "")
    ws["B3"].font = Font(name="Arial", size=11, color="6C6C75")
    kpis = dash["kpis"]
    col = 2; row = 5
    for i, k in enumerate(kpis):
        c0 = col + (i % 3) * 3
        r0 = row + (i // 3) * 4
        lab = ws.cell(row=r0, column=c0, value=k["label"])
        lab.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        lab.fill = PatternFill("solid", fgColor=accent)
        lab.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        val = ws.cell(row=r0 + 1, column=c0, value=k["formula"])
        val.font = Font(name="Arial", bold=True, size=16, color="1A1A2E")
        val.fill = PatternFill("solid", fgColor="EFEEFB")
        val.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        val.number_format = FMT.get(k.get("format", "currency0"), "General")
        for rr in (r0, r0 + 1):
            for cc in range(c0, c0 + 2):
                ws.cell(row=rr, column=cc).border = BORDER
        ws.merge_cells(start_row=r0, start_column=c0, end_row=r0, end_column=c0 + 1)
        ws.merge_cells(start_row=r0 + 1, start_column=c0, end_row=r0 + 1, end_column=c0 + 1)
        ws.column_dimensions[get_column_letter(c0)].width = 18
        ws.column_dimensions[get_column_letter(c0 + 1)].width = 10
    ch = dash.get("chart")
    if ch:
        chart = BarChart(); chart.type = "col"; chart.title = ch["title"]
        chart.height = 8; chart.width = 20; chart.legend = None
        data = Reference(ws.parent[ch["sheet"]], min_col=ch["val_col"],
                         min_row=1, max_row=ch["max_row"])
        cats = Reference(ws.parent[ch["sheet"]], min_col=ch["cat_col"],
                         min_row=2, max_row=ch["max_row"])
        chart.add_data(data, titles_from_data=True); chart.set_categories(cats)
        anchor_row = row + ((len(kpis) + 2) // 3) * 4 + 1
        ws.add_chart(chart, f"B{anchor_row}")


def build_instructions(ws, lines, brand):
    ws.sheet_view.showGridLines = False
    ws["B2"] = f"{brand} — How to use this template"
    ws["B2"].font = Font(name="Arial", bold=True, size=16, color="1A1A2E")
    r = 4
    for ln in lines:
        cell = ws.cell(row=r, column=2, value=("•  " + ln) if ln else "")
        cell.font = Font(name="Arial", size=11, color="333333")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)
        ws.row_dimensions[r].height = 28
        r += 1
    ws.column_dimensions["B"].width = 14


def fit_page(ws, landscape=True):
    from openpyxl.worksheet.properties import PageSetupProperties
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins.left = ws.page_margins.right = 0.3


def build(spec, out):
    accent = spec.get("accent", "7C6AF7")
    wb = Workbook(); wb.remove(wb.active)
    wb.create_sheet("Dashboard")
    for s in spec["sheets"]:
        ws = wb.create_sheet(s["name"])
        build_sheet(ws, s, accent)
        style_header(ws, len(s["columns"]), accent)
        for col in s.get("hide_cols", []):
            ws.column_dimensions[col].hidden = True
        fit_page(ws)
    build_dashboard(wb["Dashboard"], spec["dashboard"], accent)
    fit_page(wb["Dashboard"])
    wb.create_sheet("Instructions")
    build_instructions(wb["Instructions"], spec["instructions"], spec.get("brand", "SheetForge"))
    fit_page(wb["Instructions"], landscape=False)
    wb.save(out)
    return out


# ---------- embedded demo spec (also the exact JSON contract the LLM emits) ----------
DEMO = {
  "title": "Reseller Profit & Inventory Tracker",
  "brand": "SheetForge", "accent": "7C6AF7",
  "sheets": [
    {"name": "Sales Log", "rows": 100, "hide_cols": ["L"], "columns": [
      {"header": "Date", "type": "input", "format": "date", "width": 12},
      {"header": "Order #", "type": "input", "format": "text", "width": 14},
      {"header": "Item", "type": "input", "format": "text", "width": 26},
      {"header": "Channel", "type": "input", "format": "text", "width": 12},
      {"header": "Sale Price", "type": "input", "format": "currency", "width": 13},
      {"header": "Shipping\nCharged", "type": "input", "format": "currency", "width": 13},
      {"header": "Item Cost", "type": "input", "format": "currency", "width": 12},
      {"header": "Platform\nFees", "type": "input", "format": "currency", "width": 12},
      {"header": "Shipping\nCost", "type": "input", "format": "currency", "width": 12},
      {"header": "Net Profit", "type": "formula", "format": "currency",
       "formula": "=E{row}+F{row}-G{row}-H{row}-I{row}", "width": 13},
      {"header": "Margin %", "type": "formula", "format": "percent",
       "formula": "=IFERROR(J{row}/(E{row}+F{row}),0)", "width": 11},
      {"header": "Month", "type": "formula", "format": "text",
       "formula": "=IF(A{row}=\"\",\"\",TEXT(A{row},\"mmm\"))", "width": 8},
    ]},
    {"name": "Monthly Summary", "rows": 12, "autofilter": False, "columns": [
      {"header": "Month", "type": "seed", "format": "text", "width": 10,
       "values": ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]},
      {"header": "Revenue", "type": "formula", "format": "currency", "link": True, "width": 14,
       "formula": "=SUMIFS('Sales Log'!$E$2:$E$101,'Sales Log'!$L$2:$L$101,A{row})+SUMIFS('Sales Log'!$F$2:$F$101,'Sales Log'!$L$2:$L$101,A{row})"},
      {"header": "Item Cost", "type": "formula", "format": "currency", "link": True, "width": 13,
       "formula": "=SUMIFS('Sales Log'!$G$2:$G$101,'Sales Log'!$L$2:$L$101,A{row})"},
      {"header": "Fees + Ship", "type": "formula", "format": "currency", "link": True, "width": 13,
       "formula": "=SUMIFS('Sales Log'!$H$2:$H$101,'Sales Log'!$L$2:$L$101,A{row})+SUMIFS('Sales Log'!$I$2:$I$101,'Sales Log'!$L$2:$L$101,A{row})"},
      {"header": "Net Profit", "type": "formula", "format": "currency", "width": 14,
       "formula": "=B{row}-C{row}-D{row}"},
      {"header": "Orders", "type": "formula", "format": "int", "link": True, "width": 10,
       "formula": "=COUNTIFS('Sales Log'!$L$2:$L$101,A{row})"},
    ]},
    {"name": "Inventory", "rows": 50, "columns": [
      {"header": "SKU", "type": "input", "format": "text", "width": 14},
      {"header": "Item", "type": "input", "format": "text", "width": 28},
      {"header": "Qty on Hand", "type": "input", "format": "int", "width": 12},
      {"header": "Unit Cost", "type": "input", "format": "currency", "width": 12},
      {"header": "List Price", "type": "input", "format": "currency", "width": 12},
      {"header": "Inventory Value", "type": "formula", "format": "currency", "width": 15,
       "formula": "=C{row}*D{row}"},
      {"header": "Potential Profit", "type": "formula", "format": "currency", "width": 15,
       "formula": "=(E{row}-D{row})*C{row}"},
    ]},
  ],
  "dashboard": {
    "title": "Reseller Profit Dashboard",
    "subtitle": "Enter orders in the Sales Log — every number below updates automatically.",
    "kpis": [
      {"label": "TOTAL REVENUE", "format": "currency0", "formula": "=SUM('Monthly Summary'!B2:B13)"},
      {"label": "NET PROFIT", "format": "currency0", "formula": "=SUM('Monthly Summary'!E2:E13)"},
      {"label": "PROFIT MARGIN", "format": "percent", "formula": "=IFERROR(SUM('Monthly Summary'!E2:E13)/SUM('Monthly Summary'!B2:B13),0)"},
      {"label": "ITEM COST", "format": "currency0", "formula": "=SUM('Monthly Summary'!C2:C13)"},
      {"label": "FEES + SHIPPING", "format": "currency0", "formula": "=SUM('Monthly Summary'!D2:D13)"},
      {"label": "ORDERS", "format": "int", "formula": "=SUM('Monthly Summary'!F2:F13)"},
      {"label": "AVG PROFIT / ORDER", "format": "currency0", "formula": "=IFERROR(SUM('Monthly Summary'!E2:E13)/SUM('Monthly Summary'!F2:F13),0)"},
      {"label": "INVENTORY VALUE", "format": "currency0", "formula": "=SUM(Inventory!F2:F51)"},
      {"label": "POTENTIAL PROFIT", "format": "currency0", "formula": "=SUM(Inventory!G2:G51)"},
    ],
    "chart": {"title": "Net Profit by Month", "sheet": "Monthly Summary",
              "val_col": 5, "cat_col": 1, "max_row": 13},
  },
  "instructions": [
    "Start on the Sales Log tab. Enter one row per order: date, item, channel, sale price, shipping you charged, your item cost, platform fees, and your shipping cost.",
    "Net Profit and Margin % calculate automatically — the blue cells are the only ones you type in; black and green cells are formulas, leave them alone.",
    "The Monthly Summary tab rolls every order up by month automatically. No data entry needed there.",
    "The Dashboard updates live from your entries — total revenue, net profit, margin, orders, and average profit per order.",
    "Use the Inventory tab to track unsold stock: enter quantity, unit cost, and list price to see inventory value and potential profit.",
    "To add more than 100 orders, select the last data row in Sales Log and drag it down — the formulas copy automatically.",
    "Works in Microsoft Excel, Google Sheets, and LibreOffice Calc. Currency symbol can be changed via Format ▸ Cells.",
  ],
}

if __name__ == "__main__":
    if len(sys.argv) >= 3:
        spec = json.load(open(sys.argv[1])); out = sys.argv[2]
    else:
        spec = DEMO; out = "reseller_profit_tracker.xlsx"
    print("built:", build(spec, out))
