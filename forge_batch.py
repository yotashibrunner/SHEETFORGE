"""
forge_batch.py — stamp out a whole catalog in one run.

Reads catalog.json (list of {niche, template}), and for each:
  1. forge_spec.generate_spec()         -> JSON spec   (LLM, your key)
  2. forge_build.build()                -> .xlsx       (deterministic)
  3. recalc + error scan + ghost scan   -> QA gate     (rejects errors OR ghost data)
  4. libreoffice xlsx -> PDF            -> Etsy preview
  5. listing.txt                        -> title + tags + description (LLM)

    python forge_batch.py catalog.json
"""
import os, sys, json, subprocess
import forge_spec, forge_build
from openpyxl.utils import get_column_letter

# make scripts/recalc.py importable regardless of cwd, for the QA gate
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "scripts"))
import recalc  # noqa: E402

# Windows consoles default to cp1252; the status glyphs printed below (✓/✗/—) would
# raise UnicodeEncodeError and abort the whole batch on the first rejected file.
# Force UTF-8 so a rejection just prints and the run continues to the next item.
for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding="utf-8")
    except Exception:
        pass

OUT = "catalog"


def ghost_cells(wb, spec):
    """A built workbook has ZERO input data, so every per-row derived cell on a
    record sheet MUST be blank. Any populated value is a ghost-member bug (M-001,
    "Paid in Full" on an empty row). Returns a list of offending "Sheet!Cell=value"
    descriptions. Aggregate/summary sheets (no input column) are skipped — their KPIs
    legitimately read 0."""
    bad = []
    for sdef in spec["sheets"]:
        cols = sdef["columns"]
        if not any(c["type"] == "input" for c in cols):
            continue
        gi = forge_build.gate_column_index(cols)
        if gi is None:
            continue
        gate_letter = get_column_letter(gi + 1)
        formula_cols = [i + 1 for i, c in enumerate(cols) if c["type"] == "formula"]
        ws = wb[sdef["name"]]
        for r in range(2, sdef.get("rows", 100) + 2):
            if ws[f"{gate_letter}{r}"].value not in (None, ""):
                continue  # row is in use -> derived values are legitimate
            for ci in formula_cols:
                v = ws.cell(row=r, column=ci).value
                if v not in (None, ""):
                    bad.append(f"{sdef['name']}!{get_column_letter(ci)}{r}={v!r}")
    return bad


def qa_ok(xlsx, spec):
    """Strengthened QA gate: one LibreOffice recalc, then reject on (a) any formula
    error, or (b) ghost data on empty rows. Fail-closed if the recalc can't run."""
    wb, err = recalc.recalc_values(xlsx, 120)
    if wb is None:
        print(f"  ✗ REJECTED — QA could not run: {err}")
        return False
    nerr, by_type, _ = recalc.scan_errors(wb)
    if nerr:
        print(f"  ✗ REJECTED — {nerr} formula errors {dict(by_type)}.")
        return False
    ghosts = ghost_cells(wb, spec)
    if ghosts:
        shown = ", ".join(ghosts[:4]) + (" ..." if len(ghosts) > 4 else "")
        print(f"  ✗ REJECTED — ghost data on {len(ghosts)} empty-row cell(s): {shown}")
        return False
    return True

def make_pdf(xlsx, outdir):
    subprocess.run(["soffice", "--headless", "--convert-to", "pdf",
                    "--outdir", outdir, xlsx], capture_output=True)

def listing_copy(niche, template, title):
    prompt = (f"Write an Etsy listing for a spreadsheet template.\n"
              f"Niche: {niche}. Template: {template}.\n"
              f"Return JSON: {{\"title\": <=140 chars SEO title, "
              f"\"tags\": [13 etsy tags <=20 chars], \"description\": 120-word description}}.")
    body = json.dumps({"model": forge_spec.MODEL, "max_tokens": 1200,
        "messages": [{"role": "user", "content": prompt}]}).encode()
    import urllib.request
    req = urllib.request.Request("https://api.anthropic.com/v1/messages", data=body,
        headers={"content-type": "application/json",
                 "x-api-key": os.environ["ANTHROPIC_API_KEY"],
                 "anthropic-version": "2023-06-01"})
    data = json.loads(urllib.request.urlopen(req).read())
    return "".join(b["text"] for b in data["content"] if b["type"] == "text")

def slug(s):
    return "".join(c if c.isalnum() else "_" for c in s.lower()).strip("_")

def main(catalog_file):
    os.makedirs(OUT, exist_ok=True)
    items = json.load(open(catalog_file))
    for it in items:
        name = slug(it["template"])
        print(f"\n=== {it['template']} ({it['niche']}) ===")
        spec = forge_spec.generate_spec(it["niche"], it["template"])
        xlsx = os.path.join(OUT, name + ".xlsx")
        forge_build.build(spec, xlsx)
        if not qa_ok(xlsx, spec):
            print("  Skipping listing.")
            continue
        print("  ✓ clean")
        make_pdf(xlsx, OUT)
        open(os.path.join(OUT, name + "_listing.txt"), "w").write(
            listing_copy(it["niche"], it["template"], spec["title"]))
    print("\nCatalog ready in ./" + OUT)

if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else "catalog.json")
