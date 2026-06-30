"""
scripts/recalc.py — the QA gate.

forge_build.py writes formulas with openpyxl, which stores NO cached results.
So a clean Python read can't tell a working sheet from one full of #REF!/#DIV/0!.
This script forces a real recalculation through LibreOffice (headless), then scans
every cell for Excel error values and reports the count.

Contract (what forge_batch.py depends on):
    python scripts/recalc.py <file.xlsx> [timeout_seconds]
  -> prints ONE JSON object as the LAST stdout line, with key "total_errors"
     (0 = clean, >0 = that many error cells, -1 = could not verify).
  Diagnostics go to stderr so they never pollute the JSON line.

Fail-closed: if LibreOffice is missing or the recalc can't complete, total_errors
is -1 (NOT 0), so forge_batch rejects the file rather than shipping it unverified.
Point this at a LibreOffice install via the SOFFICE_BIN env var if it isn't on PATH.
"""
import json
import os
import shutil
import subprocess
import sys
import tempfile

# Excel error sentinels LibreOffice writes back as cached string values.
ERROR_VALUES = {
    "#NULL!", "#DIV/0!", "#VALUE!", "#REF!", "#NAME?", "#NUM!", "#N/A",
    "#SPILL!", "#CALC!", "#GETTING_DATA", "#FIELD!", "#BLOCKED!", "#CONNECT!",
}

# Common LibreOffice locations, tried after PATH / $SOFFICE_BIN.
_CANDIDATES = [
    r"C:\Program Files\LibreOffice\program\soffice.exe",
    r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
    "/Applications/LibreOffice.app/Contents/MacOS/soffice",
    "/usr/bin/soffice",
    "/usr/local/bin/soffice",
    "/snap/bin/libreoffice",
]


def _err(msg):
    print(f"[recalc] {msg}", file=sys.stderr)


def find_soffice():
    override = os.environ.get("SOFFICE_BIN")
    if override and os.path.exists(override):
        return override
    for name in ("soffice", "soffice.exe", "libreoffice"):
        hit = shutil.which(name)
        if hit:
            return hit
    for path in _CANDIDATES:
        if os.path.exists(path):
            return path
    return None


# LibreOffice user-profile override that forces formulas to ALWAYS recalculate on
# load. Without this, a headless convert keeps the (empty) cached values openpyxl
# wrote, so every formula reads back blank and even #DIV/0!/#REF! go undetected.
# OOXMLRecalcMode/ODFRecalcMode: 0 = always recalc, 1 = never, 2 = prompt.
RECALC_XCU = """<?xml version="1.0" encoding="UTF-8"?>
<oor:items xmlns:oor="http://openoffice.org/2001/registry" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
 <item oor:path="/org.openoffice.Office.Calc/Formula/Load">
  <prop oor:name="OOXMLRecalcMode" oor:op="fuse"><value>0</value></prop>
 </item>
 <item oor:path="/org.openoffice.Office.Calc/Formula/Load">
  <prop oor:name="ODFRecalcMode" oor:op="fuse"><value>0</value></prop>
 </item>
</oor:items>
"""


def seed_recalc_profile(profile_dir):
    """Pre-create a LibreOffice user profile that recalculates on load."""
    user_dir = os.path.join(profile_dir, "user")
    os.makedirs(user_dir, exist_ok=True)
    with open(os.path.join(user_dir, "registrymodifications.xcu"), "w", encoding="utf-8") as f:
        f.write(RECALC_XCU)


def force_recalc_flag(src, dst):
    """Copy the workbook to dst with fullCalcOnLoad set — a second nudge to recalc on
    open, alongside the always-recalc profile. Falls back to a plain copy."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(src)
        try:
            wb.calculation.fullCalcOnLoad = True
        except Exception:
            from openpyxl.workbook.properties import CalcProperties
            wb.calculation = CalcProperties(fullCalcOnLoad=True)
        wb.save(dst)
    except Exception as e:
        _err(f"could not set fullCalcOnLoad ({e}); converting original as-is")
        shutil.copyfile(src, dst)


def recalc_values(path, timeout=120):
    """Recalculate `path` through LibreOffice and return its computed values.

    Returns (data_only openpyxl Workbook, "") on success, or (None, error_message).
    Reusable by callers that need the recomputed cells (e.g. forge_batch's ghost-data
    check), not just an error count. The workbook is read fully into memory, so the
    temp recalc dir is cleaned up before returning."""
    soffice = find_soffice()
    if not soffice:
        return None, "LibreOffice not found; set SOFFICE_BIN or install it"

    tmp = tempfile.mkdtemp(prefix="recalc_")
    try:
        staged = os.path.join(tmp, "in.xlsx")
        force_recalc_flag(path, staged)

        # Headless convert -> xlsx through a fresh, always-recalc profile. LO loads,
        # recomputes every formula, and writes the results (incl. error cells) out.
        # Output MUST go to a separate dir: converting in.xlsx into its own folder
        # collides with the input and LibreOffice silently fails to store the result.
        outdir = os.path.join(tmp, "out")
        os.makedirs(outdir, exist_ok=True)
        profile_dir = os.path.join(tmp, "profile")
        seed_recalc_profile(profile_dir)
        profile_url = "file:///" + profile_dir.replace("\\", "/")
        cmd = [soffice, "--headless", "--nologo", "--nofirststartwizard",
               f"-env:UserInstallation={profile_url}",
               "--convert-to", "xlsx:Calc MS Excel 2007 XML",
               "--outdir", outdir, staged]
        try:
            proc = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
        except subprocess.TimeoutExpired:
            return None, f"LibreOffice recalc timed out after {timeout}s"

        out_xlsx = os.path.join(outdir, "in.xlsx")
        if not os.path.exists(out_xlsx):
            return None, (f"LibreOffice produced no output (rc={proc.returncode}): "
                          f"{proc.stderr.strip()[:200]}")

        import openpyxl
        return openpyxl.load_workbook(out_xlsx, data_only=True), ""
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def scan_errors(wb):
    """Count Excel error sentinels across a recalculated workbook.
    Returns (total_errors, {error_value: count}, cells_scanned)."""
    by_type, scanned = {}, 0
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for v in row:
                scanned += 1
                if isinstance(v, str) and v in ERROR_VALUES:
                    by_type[v] = by_type.get(v, 0) + 1
    return sum(by_type.values()), by_type, scanned


def recalc(path, timeout):
    wb, err = recalc_values(path, timeout)
    if wb is None:
        return {"file": path, "total_errors": -1, "error": err}
    total, by_type, scanned = scan_errors(wb)
    return {"file": path, "total_errors": total, "errors_by_type": by_type,
            "cells_scanned": scanned, "recalc_engine": os.path.basename(find_soffice())}


def main():
    if len(sys.argv) < 2:
        print(json.dumps({"total_errors": -1, "error": "usage: recalc.py <file.xlsx> [timeout]"}))
        return 1
    path = sys.argv[1]
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 60
    if not os.path.exists(path):
        print(json.dumps({"file": path, "total_errors": -1, "error": "file not found"}))
        return 1

    result = recalc(path, timeout)
    # LAST stdout line MUST be the JSON object (forge_batch reads splitlines()[-1]).
    print(json.dumps(result))
    return 0 if result.get("total_errors") == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
